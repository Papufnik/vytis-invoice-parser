"""
Mary Jane's Invoice Scanner -- unpacked from vytis-invoice-parser-main.zip
2026-08-18 and fixed per Sween's real-world feedback after using this for
a while: it works, but three things needed real fixes, not a rewrite.

1. SHOPIFY EXPORT REMOVED ENTIRELY. Sween confirmed the store doesn't use
   Shopify anymore -- the whole SHOPIFY_COLUMNS block, the parent/child
   variant-grouping logic, the second download button, and the "Shopify"
   half of the emailed attachments are gone.

2. BARCODE BUG, ROOT-CAUSED: the original app wrapped barcodes in Excel's
   `="12345678"` formula trick to stop Excel from mangling long digit
   strings into scientific notation on open. That trick IS correct
   Excel behavior -- but Sween confirmed this file sometimes goes
   straight to Toast's upload without ever being opened in Excel first
   ("I try to open it to check it but sometimes I get too lazy"). When
   that happens, nothing evaluates the formula, so Toast (or anything
   else reading the raw file) sees the literal text `="12345678"`,
   quotes and all -- exactly the "useless barcode" Sween described.
   There is no single CSV trick that's correct both ways at once --
   formula evaluation only happens inside a spreadsheet app opening the
   file, never when a plain CSV reader or an importer reads the same
   bytes directly.
   FIX: export as a real .xlsx file instead of .csv, with the barcode
   column's cell format explicitly set to Text (openpyxl
   number_format = '@'). This isn't a workaround -- it's what Toast's own
   official Item Library template does (confirmed against Toast's
   support docs: "Toast Retail: Build Your Import Template" explicitly
   ships an .xlsx template and warns about barcodes losing leading zeros
   if not kept as text). A genuine XLSX text-formatted cell displays
   correctly if opened in Excel AND reads back as the correct plain
   string if opened by anything else (Toast's importer, pandas,
   openpyxl) -- no formula evaluation dependency either way. Barcode
   values are also now forced to string dtype immediately after Gemini's
   CSV response is parsed, so pandas never silently upcasts a
   digit-only barcode column to int/float and strips a leading zero
   before export even happens -- the same leading-zero risk Toast's own
   docs flag independently of the scientific-notation issue.

3. EMAIL SENDING -- NOT independently verifiable from here. Sween
   reported it "just kinda never sends" the last couple months, no
   specific error noted. The SMTP logic itself (mail.smtp2go.com:2525,
   STARTTLS, login, sendmail) is structurally correct for SMTP2GO's
   documented settings, and the existing try/except already surfaces
   the real exception via st.error() rather than failing silently -- so
   if this is still broken after this fix, the on-screen error message
   the next time the button is clicked is the fastest real diagnostic,
   not another guess. Likely real-world causes, unverified: an expired
   SMTP2GO app password, a suspended/rate-limited SMTP2GO account, or
   st.secrets values lost on a redeploy. Cleaned up to attach the one
   remaining Toast file (was attaching two files, Toast + Shopify)
   and updated the button label/body text accordingly -- did not touch
   the underlying SMTP call itself since there's no diagnosed bug in it.
   UPDATE 2026-08-18, after real use: sending itself worked, but Sween
   reported attachments arriving with no filename and no recognizable
   type. Root cause found -- see the attachment-building loop's own
   comment near MIMEApplication for the fix (was MIMEBase +
   "application/octet-stream", a generic untyped placeholder that never
   told the recipient's mail client this was an .xlsx file at all).

4. PURCHASING & RECEIVING EXPORT ADDED (separate follow-up request, same
   day). Sween's actual bottleneck wasn't the scan -- it was Toast's own
   receiving screen, where every item has to be searched for and added
   one at a time. Toast Retail has a bulk "Receive via invoice -> Import
   file" feature for exactly this, but it needs a DIFFERENT column
   template than the Item Library export above (verified against Toast's
   own support article, "Toast Retail: Import Invoices", 2026-08-18 --
   not guessed): Supplier Item ID, Item Name, and Item Quantity required;
   Receiving Unit Net Cost, Extension Cost, and Barcode recognized.
   Quantity wasn't extracted at all before this -- added to the Gemini
   prompt as its own field, same handwritten-correction-takes-priority
   pattern already used for cost/price. See build_receiving_xlsx()'s own
   docstring for the exact column mapping and what was deliberately left
   out (Price, Receiving Unit, PLU) and why. Sween confirmed invoices
   consistently show clear quantities and generally carry a matching
   order number back to his Order List -- the SEPARATE "auto-mark
   received on the Order List" idea from this same conversation is
   explicitly NOT built here; it needs its own pass since it means
   writing to a live shared Google Sheet automatically, a bigger and
   riskier change than a second export file.

5. CATALOG DUPLICATE CHECK, ITEM LIBRARY EXPORT ONLY (2026-08-18). The
   real risk with an invoice scanner that always suggests "new item" is
   silently re-creating an item Toast already has -- Sween's catalog
   currently runs ~14k items. This app now checks every extracted line
   against a live mirror of that catalog (see load_catalog_lookup() and
   match_catalog() below, and push_catalog_lookup_sheet.py's header for
   how the mirror gets there -- this app has no direct access to the
   local warehouse.sqlite it's built from). A likely-existing item gets
   its "Add to Item Library?" checkbox unchecked by default -- excluded
   from that export -- with the match reason shown so it's a decision a
   human can override, not a silent drop. This is DELIBERATELY SCOPED TO
   THE ITEM LIBRARY EXPORT ONLY: the Receiving export always includes
   every scanned line regardless of the checkbox, because a restock of
   an item Toast already has still needs to be received -- only genuinely
   NEW items need to go through the Item Library import at all. If the
   catalog mirror isn't reachable (secrets not configured yet, or a
   fetch error), this whole check silently no-ops and every item behaves
   exactly as before -- it's a helpful extra, never a dependency the
   rest of the app should break over.
"""

import streamlit as st
import google.generativeai as genai
import pandas as pd
import io
import smtplib
import re
import difflib
import gspread
from google.oauth2.service_account import Credentials
from PIL import Image
from openpyxl.utils import get_column_letter
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from datetime import datetime


# --- PAGE CONFIG & MOBILE UI ---
st.set_page_config(page_title="Invoice Scanner", page_icon="🧾", layout="wide")

# This CSS hides the Streamlit top menu, footer, and styles the buttons
st.markdown("""
    <style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    .stButton>button { height: 3.5em; font-size: 18px; font-weight: bold; border-radius: 8px; }
    [data-testid="stMetricValue"] { font-size: 1.8rem; color: #1E88E5; }
    </style>
""", unsafe_allow_html=True)

# --- PASSWORD GATE ---
try:
    APP_PASSWORD = st.secrets["APP_PASSWORD"]
except KeyError:
    st.error("⚠️ App password not found in server secrets. Please contact Administrator.")
    st.stop()

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

if not st.session_state.authenticated:
    st.title("Mary Jane's Scanner")
    entered = st.text_input("Enter password to continue:", type="password")
    if st.button("Login", use_container_width=True):
        if entered == APP_PASSWORD:
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error("Incorrect password.")
    st.stop()
# --- END PASSWORD GATE ---

# --- SIDEBAR CONFIG ---
st.sidebar.title("⚙️ Instructions")
st.sidebar.markdown("""
1. Upload vendor invoice images.
2. Add context if needed (e.g., 'All items are new', 'Markup 4x').
3. Click Extract Data.
4. Review the data on screen -- check quantities too, not just cost/price.
5. If an item looks like it's already in Toast, its **Add to Item Library?** box comes unchecked automatically -- check the note next to it and re-check the box yourself if it's actually a genuinely new item.
6. Two files come out: **Item Library** (only the checked, genuinely-new items) and **Receiving** (every scanned line, new or restock -- this is the one that goes into Toast's Purchasing & Receiving bulk import, so you don't have to search for each item by hand).
7. Email both to the Back Office, or download them directly.
""")

# --- MAIN UI ---
st.title("Mary Jane's Invoice Scanner 🧾")
st.write("Snap a picture of a vendor packing slip or invoice.")

extra_instructions = st.text_area(
    "Context / Instructions (Optional)",
    placeholder="e.g., 'All items on this invoice are new.', 'Markup is 4x instead of 3x.'",
)

uploaded_files = st.file_uploader("Upload Image(s)", type=["jpg", "jpeg", "png"], accept_multiple_files=True)


def get_system_prompt(user_instructions):
    return f"""
    You are a Retail Inventory Migration Specialist extracting data from wholesale invoice images.
    Your objective is to extract data and format it perfectly into a CSV.

    CRITICAL RULE: Output ONLY valid, raw CSV text. Do not include markdown wrappers (like ```csv).

    CSV Columns Required Exactly:
    name,pos name,category group,category,subcategory,price,cost,barcode,supplier,color,size,quantity

    Extraction & Logic Rules:
    1. Filtering: ONLY extract line items that are hand-marked, underlined, or explicitly labeled as "New".
       Prioritize user instructions regarding which items to process.
    2. Naming Convention (name column):
       - Clothing Sized: [SKU]-[Color]-[Size]
       - Clothing One-Size: [SKU]-[Color]
       - Non-Clothing: Just the [SKU]
    3. Color & Size Columns: Extract the Color and Size from the description and place them in their respective columns. If none exist, leave blank.
    4. POS Name: Copy the description verbatim.
    5. Category Mapping: Map based on item type (Accessories, Beer, BTG Wine, Clothing, Gifts, Handbags, Hats, Home, Jewelry, Snacks & Drinks, Wine Bottles). Category Group is ALWAYS "Retail".
    6. Cost & Price: Handwritten value if present; else printed unit cost. Price is Handwritten retail price, or Cost * 3. ALWAYS round the retail price to nearest dollar.
    7. Barcode: Printed UPC/barcode or leave blank. Preserve every digit exactly as printed, including leading zeros -- do not treat it as a number.
    8. Subcategory & Supplier: Use the brand name found at top of invoice.
    9. Quantity: The actual quantity received for this line item -- use a handwritten correction if present (e.g. crossed-out and rewritten count), otherwise the printed/ordered quantity. Output as a plain whole number, no units or text. Never leave blank -- if genuinely illegible, output 0 rather than guessing.

    User Additional Instructions: {user_instructions}
    """


# --- HELPER: LOGICAL SIZE SORTING ---
def get_size_rank(size_val):
    size_str = str(size_val).upper().strip()
    size_order = {
        'XXS': 1, 'XS': 2, 'S': 3, 'SMALL': 3,
        'M': 4, 'MEDIUM': 4, 'L': 5, 'LARGE': 5,
        'XL': 6, 'XXL': 7, '2XL': 7, '3XL': 8, '4XL': 9,
        'OS': 0, 'ONE SIZE': 0
    }
    if size_str in size_order:
        return size_order[size_str]

    num_match = re.search(r'\d+', size_str)
    if num_match:
        return 100 + float(num_match.group())

    return 50


# --- HELPER: BUILD BARCODE-SAFE XLSX (see module docstring, fix #2) ---
def build_toast_xlsx(df):
    """Writes df to an in-memory .xlsx (not .csv) with the barcode column's
    cell format forced to Text, so it displays correctly if a human opens
    it in Excel AND reads back correctly as a plain string if anything
    else (Toast's importer, pandas, openpyxl) reads the file directly --
    no formula-evaluation dependency either way. See module docstring."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Toast Import")
        worksheet = writer.sheets["Toast Import"]
        if "barcode" in df.columns:
            col_idx = list(df.columns).index("barcode") + 1  # openpyxl columns are 1-indexed
            col_letter = get_column_letter(col_idx)
            for row_num in range(2, len(df) + 2):  # row 1 is the header
                worksheet[f"{col_letter}{row_num}"].number_format = "@"
    return output.getvalue()


def build_receiving_xlsx(edited_df):
    """Second export, added 2026-08-18: Toast Retail's Purchasing & Receiving
    'Import file' format (Retail > Purchasing > Purchasing & Receiving >
    Receive via invoice > Import file) -- a DIFFERENT template from the
    Item Library import build_toast_xlsx() produces above. Columns and
    requirements verified directly against Toast's own support article
    ("Toast Retail: Import Invoices", support.toasttab.com) on 2026-08-18,
    not guessed:
      - Required: Supplier Item ID, Item Name, Item Quantity
      - Recognized optional: Receiving Unit, Qty/Receiving Unit,
        Receiving Unit Net Cost, Extension Cost, Barcode, PLU, Price (Retail)

    Column mapping decisions, stated explicitly rather than silently:
      - UPDATE 2026-08-19, after real use: Sween confirmed Toast's actual
        receiving screen matched ZERO of these against existing catalog
        items -- the original theory below (Supplier Item ID as the
        primary match key) doesn't hold up against how Toast's receiving
        import actually matches in practice. What it really keys off is
        Item Name against the catalog's own `name` field -- and this
        app's `name` (the SKU-style field, e.g. "P1015-Beige-L") IS
        already confirmed (see match_catalog()/extract_base_sku() above)
        to be the same convention Toast's own `name` field uses for
        anything entered through this app. So `name` now goes in BOTH
        Item Name and Supplier Item ID -- `pos name` (verbatim
        description) is no longer used in this export at all, since it
        was never what Toast was actually matching against.
      - Supplier Item ID <- `name`, same SKU-style value as Item Name.
        Kept populated too since Toast's own docs describe it as a
        secondary/fallback signal -- redundant with Item Name now, but
        harmless, and Toast's own review screen still lets you manually
        link anything that doesn't match automatically either way.
      - Item Name <- `name` (the SKU-style field) -- see UPDATE above for
        why this replaced `pos name`.
      - Item Quantity <- new `quantity` field (see the Gemini prompt above).
      - Receiving Unit Net Cost <- `cost`.
      - Extension Cost <- cost * quantity, computed here rather than left
        for Toast to derive, so the number is visible and checkable before
        upload.
      - Barcode <- same barcode column, same text-format safety fix as
        build_toast_xlsx() above (identical bug, identical fix).
      - Receiving Unit, Qty/Receiving Unit, PLU: deliberately left OUT.
        Nothing in the current extraction reliably determines these
        (e.g. whether an item is received by the each vs. by the case),
        and Toast's own import treats missing optional columns as fine --
        better to omit than guess.
      - Price (Retail) <- `price`, INCLUDED despite Toast's own docs
        saying the automated import "will not update" retail price on
        items it already recognizes -- Sween's correct catch (2026-08-18)
        that this still matters for genuinely NEW items: Toast's own
        receiving flow lets you create a new item on the spot when
        nothing matches, and that manual popup asks for a price right
        then. Having it already sitting in this same file means not
        needing to cross-reference the Item Library file at that moment,
        even though the bulk-import mechanism itself won't read it back
        onto an existing item. Reference value for the human, not a
        live-updated field for Toast's matcher -- worth knowing the
        difference, not silently treating it as equivalent to the other
        columns here.
    """
    receiving_df = pd.DataFrame({
        "Supplier Item ID": edited_df["name"],
        "Item Name": edited_df["name"],
        "Item Quantity": pd.to_numeric(edited_df["quantity"], errors="coerce").fillna(0).astype(int),
        "Receiving Unit Net Cost": pd.to_numeric(edited_df["cost"], errors="coerce"),
        "Barcode": edited_df["barcode"],
        "Price (Retail)": pd.to_numeric(edited_df["price"], errors="coerce"),
    })
    receiving_df["Extension Cost"] = (receiving_df["Item Quantity"] * receiving_df["Receiving Unit Net Cost"]).round(2)
    receiving_df = receiving_df[["Supplier Item ID", "Item Name", "Item Quantity", "Receiving Unit Net Cost", "Extension Cost", "Barcode", "Price (Retail)"]]

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        receiving_df.to_excel(writer, index=False, sheet_name="Receiving Import")
        worksheet = writer.sheets["Receiving Import"]
        col_idx = list(receiving_df.columns).index("Barcode") + 1
        col_letter = get_column_letter(col_idx)
        for row_num in range(2, len(receiving_df) + 2):
            worksheet[f"{col_letter}{row_num}"].number_format = "@"
    return output.getvalue()


# --- CATALOG DUPLICATE CHECK (feature 5, see module docstring) ---
CATALOG_SHEET_SCOPES = ["https://www.googleapis.com/auth/spreadsheets.readonly"]


def _norm(s):
    return str(s or "").strip().lower()


@st.cache_data(ttl=3600, show_spinner=False)
def load_catalog_lookup():
    """Reads the "MaryJanes_Catalog_Lookup" Google Sheet -- a mirror of
    the real Toast catalog, pushed from the local db/warehouse.sqlite by
    push_catalog_lookup_sheet.py (see that script's header for the full
    why/how). Cached 1 hour per session so a batch of invoices doesn't
    re-hit the Sheets API on every st.rerun().

    Returns None (never raises) if either secret below is missing or the
    fetch fails for any reason -- this check is a helpful extra, not
    something the rest of the app should ever depend on or break over.
    Required st.secrets, added in the Streamlit Cloud app settings:
      - CATALOG_LOOKUP_SHEET_ID: the sheet's ID (from its URL)
      - [gcp_service_account]: the same service account JSON already
        used by this project's other scripts, as a TOML table
    """
    try:
        sheet_id = st.secrets["CATALOG_LOOKUP_SHEET_ID"]
        sa_info = dict(st.secrets["gcp_service_account"])
    except KeyError:
        return None

    try:
        creds = Credentials.from_service_account_info(sa_info, scopes=CATALOG_SHEET_SCOPES)
        gc = gspread.authorize(creds)
        ws = gc.open_by_key(sheet_id).worksheet("Catalog")
        cat_df = pd.DataFrame(ws.get_all_records())
        for col in ["name", "pos_name", "barcode", "supplier"]:
            if col not in cat_df.columns:
                cat_df[col] = ""
            cat_df[col] = cat_df[col].astype(str)
        return cat_df
    except Exception:
        return None


def extract_base_sku(row):
    """Strips the Color/Size suffix off an extracted item's `name`, using
    THIS SAME ROW's own separately-extracted `color`/`size` fields --
    never guessed -- per the exact convention get_system_prompt() tells
    Gemini to build `name` with:
      Clothing Sized:    [SKU]-[Color]-[Size]
      Clothing One-Size: [SKU]-[Color]
      Non-Clothing:      [SKU]              (nothing to strip)
    Added 2026-08-18 per Sween's point: since this app is the one that
    defines that naming convention, knowing it exactly is a stronger
    signal than a blind full-string comparison -- it also correctly
    resolves a genuinely new color/size of an EXISTING product back to
    the same base SKU the catalog already has, which a raw exact-match
    on the full `name` string would miss.
    """
    name = str(row.get("name", "") or "").strip()
    color = str(row.get("color", "") or "").strip()
    size = str(row.get("size", "") or "").strip()
    suffix = "-".join(p for p in (color, size) if p)
    if suffix and name.lower().endswith(("-" + suffix).lower()):
        return name[: -(len(suffix) + 1)]
    return name


def match_catalog(df, cat_df):
    """Flags each extracted invoice line against the catalog mirror.
    Four passes, strongest signal first -- see push_catalog_lookup_
    sheet.py's header for why the catalog's own `name` field doesn't
    follow one single convention (some items are this app's own
    "SKU-Color-Size" style, some are legacy "SKU : Description" strings
    entered a different way), so no one exact-match rule catches
    everything alone:

      1. Barcode -- exact string match on a non-blank barcode both
         sides. Strongest possible signal (same physical UPC).
      2. Base SKU -- extract_base_sku() strips this row's own Color/Size
         suffix off `name` per the app's known naming formula, then
         compares that against the catalog `name` field's own portion
         before " : " (the SKU-style prefix Toast's export shows for
         items entered this same way -- e.g. catalog name
         "CWAH1275-Asst : Vintage Washed Distressed Mountain Cap" has
         base SKU "CWAH1275-Asst"). Catches a new color/size of an
         existing product that a raw full-string match would miss.
      3. Full name/pos_name -- case-insensitive exact match of this
         app's `name` OR `pos name` against the catalog's `name` OR
         `pos_name`, for whatever base-SKU splitting doesn't cover.
      4. Fuzzy description -- same `supplier`, difflib similarity ratio
         >= 0.90 between `pos name` values. Lower confidence (a real new
         color/size variant of an existing product can still score high
         here) -- flagged as "possible" but NOT auto-excluded, left for
         a human glance rather than a silent drop.

    Returns a copy of df with two new columns:
      _dup_confidence: "" | "possible" | "likely"
      _dup_detail: human-readable reason, or "" if no match.
    """
    barcode_map, name_map, sku_map, by_supplier = {}, {}, {}, {}
    for _, r in cat_df.iterrows():
        bc = _norm(r.get("barcode", ""))
        if bc:
            barcode_map.setdefault(bc, r)
        cat_name = str(r.get("name", "") or "")
        for key in (_norm(cat_name), _norm(r.get("pos_name", ""))):
            if key:
                name_map.setdefault(key, r)
        cat_sku = _norm(cat_name.split(" : ")[0])
        if cat_sku:
            sku_map.setdefault(cat_sku, r)
        by_supplier.setdefault(_norm(r.get("supplier", "")), []).append(r)

    confidences, details = [], []
    for _, row in df.iterrows():
        bc, nm, pnm, sup = (
            _norm(row.get("barcode", "")), _norm(row.get("name", "")),
            _norm(row.get("pos name", "")), _norm(row.get("supplier", "")),
        )
        base_sku = _norm(extract_base_sku(row))

        conf, detail = "", ""
        if bc and bc in barcode_map:
            conf, detail = "likely", f'barcode already in catalog ("{barcode_map[bc].get("name", "")}")'
        elif base_sku and base_sku in sku_map:
            conf, detail = "likely", f'SKU "{extract_base_sku(row)}" matches existing item "{sku_map[base_sku].get("name", "")}"'
        elif nm and nm in name_map:
            conf, detail = "likely", f'matches existing item "{name_map[nm].get("name", "")}"'
        elif pnm and pnm in name_map:
            conf, detail = "likely", f'matches existing item "{name_map[pnm].get("name", "")}"'
        else:
            best_ratio, best_row = 0.0, None
            for cand in by_supplier.get(sup, []):
                ratio = difflib.SequenceMatcher(None, pnm, _norm(cand.get("pos_name", ""))).ratio()
                if ratio > best_ratio:
                    best_ratio, best_row = ratio, cand
            if best_ratio >= 0.90 and best_row is not None:
                conf, detail = "possible", f'similar to existing "{best_row.get("pos_name", "")}" ({best_ratio:.0%} match)'

        confidences.append(conf)
        details.append(detail)

    out = df.copy()
    out["_dup_confidence"] = confidences
    out["_dup_detail"] = details
    return out


# --- PROCESSING ---
if uploaded_files:
    file_names = [f.name for f in uploaded_files]
    if "current_files" not in st.session_state or st.session_state.current_files != file_names:

        if st.button("✨ Extract Data", use_container_width=True, type="primary"):
            try:
                backend_api_key = st.secrets["GEMINI_API_KEY"]
            except KeyError:
                st.error("⚠️ System Error: Gemini API Key missing from server secrets. Contact Administrator.")
                st.stop()

            with st.spinner("Analyzing invoices with Gemini Pro..."):
                try:
                    genai.configure(api_key=backend_api_key)
                    model = genai.GenerativeModel('gemini-pro-latest')

                    prompt = get_system_prompt(extra_instructions)
                    images = [Image.open(file) for file in uploaded_files]
                    inputs = [prompt] + images

                    response = model.generate_content(inputs)

                    raw_csv = response.text.strip()
                    if raw_csv.startswith("```csv"): raw_csv = raw_csv[6:]
                    if raw_csv.startswith("```"): raw_csv = raw_csv[3:]
                    if raw_csv.endswith("```"): raw_csv = raw_csv[:-3]
                    raw_csv = raw_csv.strip()

                    df = pd.read_csv(io.StringIO(raw_csv))

                    # Force barcode to string dtype immediately -- otherwise
                    # a digit-only barcode column can get silently upcast to
                    # int64/float64 by read_csv, which can drop a leading
                    # zero before the sheet is even built (see fix #2 in
                    # the module docstring -- this is a distinct risk from
                    # the scientific-notation display bug, flagged
                    # independently in Toast's own import docs).
                    if "barcode" in df.columns:
                        df["barcode"] = df["barcode"].astype(str).replace("nan", "")

                    # Apply Custom Logical Sorting
                    if 'size' in df.columns and 'color' in df.columns:
                        df['_size_rank'] = df['size'].apply(get_size_rank)
                        df = df.sort_values(by=['pos name', 'color', '_size_rank'], na_position='first')
                        df = df.drop(columns=['_size_rank'])

                    st.session_state.invoice_data = df
                    st.session_state.current_files = file_names
                    st.rerun()

                except Exception as e:
                    st.error(f"❌ An error occurred: {str(e)}")

# --- DISPLAY & EXPORT ---
if "invoice_data" in st.session_state and not st.session_state.invoice_data.empty:
    st.success("✅ Extraction Complete! Review the data below.")

    # --- CATALOG DUPLICATE CHECK (feature 5, see module docstring) ---
    # Runs BEFORE the data_editor so the resulting checkbox is itself
    # editable alongside any other hand-correction, and so a user can see
    # + override the auto-exclusion in the same pass as reviewing the rest.
    base_df = st.session_state.invoice_data.copy()
    catalog_df = load_catalog_lookup()
    if catalog_df is not None and not catalog_df.empty:
        matched_df = match_catalog(base_df, catalog_df)
        add_to_library = matched_df["_dup_confidence"] != "likely"
        note = matched_df["_dup_detail"]
        base_df.insert(0, "Catalog Match Note", note)
        base_df.insert(0, "Add to Item Library?", add_to_library)

        n_likely = int((matched_df["_dup_confidence"] == "likely").sum())
        n_possible = int((matched_df["_dup_confidence"] == "possible").sum())
        if n_likely or n_possible:
            st.warning(
                f"🔎 Catalog check: {n_likely} item(s) look like they're already in Toast "
                f"(unchecked below, excluded from the Item Library file) and {n_possible} "
                f"look similar to an existing item (left checked -- worth a glance). "
                f"This never changes the Receiving file -- every scanned line still gets received."
            )
    else:
        base_df.insert(0, "Catalog Match Note", "")
        base_df.insert(0, "Add to Item Library?", True)

    edited_export_df = st.data_editor(base_df, use_container_width=True, hide_index=True)

    # Clean data to prevent missing columns from breaking export
    for col in ['color', 'size', 'barcode', 'quantity']:
        if col not in edited_export_df.columns:
            edited_export_df[col] = '' if col != 'quantity' else 0

    edited_export_df['color'] = edited_export_df['color'].fillna('')
    edited_export_df['size'] = edited_export_df['size'].fillna('')
    edited_export_df['barcode'] = edited_export_df['barcode'].astype(str).replace('nan', '').fillna('')
    edited_export_df['quantity'] = pd.to_numeric(edited_export_df['quantity'], errors='coerce').fillna(0).astype(int)

    # --- DYNAMIC FILE NAMING ---
    try:
        brand_name = str(edited_export_df['subcategory'].iloc[0]).strip()
        if not brand_name or brand_name.lower() == "nan": brand_name = "Invoice"
    except Exception:
        brand_name = "Invoice"

    date_str = datetime.now().strftime("%m%d%Y")
    toast_filename = f"{brand_name} {date_str} - Item Library.xlsx"
    receiving_filename = f"{brand_name} {date_str} - Receiving.xlsx"

    # --- TOAST EXPORTS ---
    # Two DIFFERENT Toast imports, added 2026-08-18 -- see build_receiving_xlsx's
    # docstring for exactly why these are separate files, not one:
    #   1. Item Library import -- creates/updates the catalog item itself
    #      (name, category, price, cost, barcode, supplier). Only needed for
    #      genuinely NEW items -- Toast already has everything about an
    #      existing item.
    #   2. Purchasing & Receiving import -- logs THIS shipment (quantity
    #      received, net cost paid) against an item already in the catalog.
    #      Needed every time, new item or restock, since it's what actually
    #      updates on-hand inventory count and clears the item-by-item
    #      manual search Sween described as the real bottleneck.
    # Item Library respects the "Add to Item Library?" checkbox (feature 5,
    # see module docstring) -- items flagged as likely already in Toast are
    # excluded here unless a human re-checks the box. Receiving does NOT
    # filter on this at all: build_receiving_xlsx() gets the full
    # edited_export_df, every scanned line, because a restock of an
    # existing item still needs to be received -- only the Item Library
    # import (which would try to CREATE the item) needs the exclusion.
    if "Add to Item Library?" in edited_export_df.columns:
        item_library_rows = edited_export_df[edited_export_df["Add to Item Library?"]]
    else:
        item_library_rows = edited_export_df
    toast_output = item_library_rows[['name', 'pos name', 'category group', 'category', 'subcategory', 'price', 'cost', 'barcode', 'supplier']].copy()
    # Added 2026-08-19: Sween's real-world test found Toast's Receiving
    # import can't match ANY item back to its catalog record if that
    # record's own `supplier item id` field was never set -- and it never
    # was, for anything created through this Item Library file, because
    # this column didn't used to be here. Backfilled once for existing
    # items via a separate one-time fix (MaryJanes_SupplierItemID_
    # Backfill_ToastImport.csv); this closes the gap going forward so
    # every NEW item this app creates gets it saved at creation time.
    # Same `name` value already used as Supplier Item ID in the Receiving
    # export -- one consistent SKU-style identifier across both files.
    toast_output.insert(1, "supplier item id", toast_output["name"])
    toast_bytes = build_toast_xlsx(toast_output)
    receiving_bytes = build_receiving_xlsx(edited_export_df)

    st.divider()

    # --- EXPORT BUTTONS ---
    col1, col2, col3 = st.columns([2, 1, 1])

    with col1:
        if st.button("📤 Email Both Files to Back Office", use_container_width=True):
            try:
                sender = st.secrets["SENDER_EMAIL"]
                recipient = st.secrets["RECIPIENT_EMAIL"]
                sender_pwd = st.secrets["SENDER_APP_PASSWORD"]

                msg = MIMEMultipart()
                msg["From"] = sender
                msg["To"] = recipient
                msg["Subject"] = f"Invoice Upload - {brand_name}"
                msg.attach(MIMEText(
                    f"Attached for {brand_name}: the Item Library file (for any new items) "
                    f"and the Receiving file (to log this shipment's quantity/cost -- use this "
                    f"one every time, new items or restocks).",
                    "plain",
                ))

                # Fixed 2026-08-18 -- Sween reported attachments arriving
                # with no name and no recognizable file type. Root cause:
                # MIMEBase("application", "octet-stream") declares every
                # attachment as generic untyped binary data, and only the
                # Content-Disposition header carried a filename -- some
                # mail clients read the filename from Content-Type's own
                # `name` parameter instead (or in addition), so a client
                # that checks there first found nothing. MIMEApplication
                # with the real xlsx subtype + Name= sets BOTH: the
                # correct application/vnd.openxmlformats-officedocument.
                # spreadsheetml.sheet type (same one already used for the
                # download buttons below) AND the filename in Content-Type,
                # on top of the Content-Disposition filename already being
                # set explicitly -- covers whichever header a given client
                # actually reads.
                for file_bytes, filename in ((toast_bytes, toast_filename), (receiving_bytes, receiving_filename)):
                    part = MIMEApplication(
                        file_bytes,
                        _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        Name=filename,
                    )
                    part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
                    msg.attach(part)

                with smtplib.SMTP("mail.smtp2go.com", 2525) as server:
                    server.starttls()
                    server.login(sender, sender_pwd)
                    server.sendmail(sender, recipient, msg.as_string())

                st.success("✅ Both files sent to Back Office successfully!")
            except Exception as e:
                st.error(f"Failed to send email: {type(e).__name__}: {e}")

    with col2:
        st.download_button(
            label="⬇️ Item Library File",
            data=toast_bytes,
            file_name=toast_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    with col3:
        st.download_button(
            label="⬇️ Receiving File",
            data=receiving_bytes,
            file_name=receiving_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    st.divider()

    if st.button("🔄 Scan a New Invoice", use_container_width=True):
        del st.session_state.invoice_data
        del st.session_state.current_files
        st.rerun()
